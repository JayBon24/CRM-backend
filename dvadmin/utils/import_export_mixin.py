# -*- coding: utf-8 -*-
import datetime
import io
import os
from urllib.parse import quote

from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.table import Table, TableStyleInfo
from rest_framework.decorators import action
from rest_framework.request import Request
from rest_framework.response import Response

from dvadmin.utils.import_export import import_to_data
from dvadmin.utils.json_response import DetailResponse, SuccessResponse
from dvadmin.utils.request_util import get_verbose_name
from dvadmin.system.tasks import async_export_data
from dvadmin.system.models import DownloadCenter


class ImportSerializerMixin:
    """
    自定义导入模板、导入功能
    """

    # 导入字段
    import_field_dict = {}

    def filter_queryset_for_import_delete(self, queryset):
        """
        导入时若启用「删除未在 Excel 中的记录」，会先过滤出可被删除的 queryset。
        子类可重写以排除受保护记录（如超管、当前用户）。
        默认不做过滤。
        """
        return queryset
    # 导入序列化器
    import_serializer_class = None
    # 表格表头最大宽度，默认50个字符
    export_column_width = 50

    def is_number(self, num):
        if isinstance(num, (list, dict)):
            return False
        try:
            float(num)
            return True
        except (ValueError, TypeError):
            pass
        try:
            import unicodedata
            unicodedata.numeric(num)
            return True
        except (TypeError, ValueError):
            pass
        return False

    def get_string_len(self, string):
        """
        获取字符串最大长度；支持 list/dict 等非标量（转为字符串后计算）
        """
        length = 4
        if string is None:
            return length
        if isinstance(string, list):
            string = ", ".join(str(x) for x in string)
        elif isinstance(string, dict):
            string = str(string)
        if self.is_number(string):
            return length
        try:
            for char in str(string):
                length += 2.1 if ord(char) > 256 else 1
        except (TypeError, ValueError):
            return length
        return round(length, 1) if length <= self.export_column_width else self.export_column_width

    @action(methods=['get','post'],detail=False)
    @transaction.atomic  # Django 事务,防止出错
    def import_data(self, request: Request, *args, **kwargs):
        """
        导入模板
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        assert self.import_field_dict, "'%s' 请配置对应的导出模板字段。" % self.__class__.__name__
        # 导出模板
        if request.method == "GET":
            # 示例数据
            queryset = self.filter_queryset(self.get_queryset())
            # 导出excel 表
            response = HttpResponse(content_type="application/msexcel")
            response["Access-Control-Expose-Headers"] = f"Content-Disposition"
            response[
                "Content-Disposition"
            ] = f'attachment;filename={quote(str(f"导入{get_verbose_name(queryset)}模板.xlsx"))}'
            wb = Workbook()
            ws1 = wb.create_sheet("data", 1)
            ws1.sheet_state = "hidden"
            ws = wb.active
            row = get_column_letter(len(self.import_field_dict) + 1)
            column = 10
            header_data = [
                "序号",
            ]
            validation_data_dict = {}
            validation_field_index = {}  # 记录每个验证字段在表头中的列索引
            current_header_index = 1  # 从1开始（0是序号列）
            for index, ele in enumerate(self.import_field_dict.values()):
                if isinstance(ele, dict):
                    header_title = ele.get("title")
                    header_data.append(str(header_title) if header_title is not None else "")
                    current_header_index += 1
                    choices = ele.get("choices", {})
                    if choices.get("data"):
                        data_list = []
                        data_list.extend(choices.get("data").keys())
                        validation_data_dict[header_title] = data_list
                        validation_field_index[header_title] = current_header_index
                    elif choices.get("queryset") and choices.get("values_name"):
                        queryset = choices.get("queryset")
                        values_name = choices.get("values_name")
                        # 确保 queryset 是可执行的（可能是 lazy queryset）
                        try:
                            data_list = list(queryset.values_list(values_name, flat=True).distinct())
                        except:
                            data_list = []
                        if data_list:
                            validation_data_dict[header_title] = data_list
                            validation_field_index[header_title] = current_header_index
                    # 如果没有 choices 配置，跳过下拉验证
                else:
                    header_data.append(str(ele) if ele is not None else "")
                    current_header_index += 1
            # 添加数据列到隐藏的 data sheet
            if validation_data_dict:
                ws1.append(list(validation_data_dict.keys()))
                for validation_index, (field_title, validation_data) in enumerate(validation_data_dict.items()):
                    for inx, ele in enumerate(validation_data):
                        ws1[f"{get_column_letter(validation_index + 1)}{inx + 2}"] = ele
                # 为每个有下拉选项的字段添加数据验证
                for field_title, validation_data in validation_data_dict.items():
                    if field_title in validation_field_index:
                        header_col_index = validation_field_index[field_title]
                        data_col_letter = get_column_letter(list(validation_data_dict.keys()).index(field_title) + 1)
                        dv = DataValidation(
                            type="list",
                            formula1=f"{quote_sheetname('data')}!${data_col_letter}$2:${data_col_letter}${len(validation_data) + 1}",
                            allow_blank=True,
                        )
                        ws.add_data_validation(dv)
                        header_col_letter = get_column_letter(header_col_index)
                        dv.add(f"{header_col_letter}2:{header_col_letter}1048576")
            # 插入导出模板正式数据
            df_len_max = [self.get_string_len(ele) for ele in header_data]
            ws.append(header_data)
            # 　更新列宽
            for index, width in enumerate(df_len_max):
                ws.column_dimensions[get_column_letter(index + 1)].width = width
            tab = Table(displayName="Table1", ref=f"A1:{row}{column}")  # 名称管理器
            style = TableStyleInfo(
                name="TableStyleLight11",
                showFirstColumn=True,
                showLastColumn=True,
                showRowStripes=True,
                showColumnStripes=True,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)
            wb.save(response)
            return response
        else:
            # 从excel中组织对应的数据结构，然后使用序列化器保存
            queryset = self.filter_queryset(self.get_queryset())
            # 获取多对多字段
            m2m_fields = [
                ele.name
                for ele in queryset.model._meta.get_fields()
                if hasattr(ele, "many_to_many") and ele.many_to_many == True
            ]
            import_field_dict = {'id':'更新主键(勿改)',**self.import_field_dict}
            data = import_to_data(request.data.get("url"), import_field_dict, m2m_fields)
            for ele in data:
                filter_dic = {'id':ele.get('id')}
                instance = filter_dic and queryset.filter(**filter_dic).first()
                # print(156,ele)
                serializer = self.import_serializer_class(instance, data=ele, context={"request": request})
                serializer.is_valid(raise_exception=True)
                serializer.save()
            # 可选：删除 Excel 中未出现的记录（仅当请求带 delete_missing=True 且 Excel 中有 id 列时）
            if request.data.get('delete_missing') and data:
                ids_in_excel = {e.get('id') for e in data if e.get('id') is not None}
                if ids_in_excel:
                    to_delete = queryset.exclude(id__in=ids_in_excel)
                    to_delete = self.filter_queryset_for_import_delete(to_delete)
                    deleted_count = to_delete.count()
                    to_delete.delete()
                    if deleted_count:
                        return DetailResponse(msg=f"导入成功！已删除 {deleted_count} 条未在 Excel 中的记录。")
            return DetailResponse(msg=f"导入成功！")

    @action(methods=['get'],detail=False)
    def update_template(self,request):
        try:
            queryset = self.filter_queryset(self.get_queryset())
        except Exception as e:
            return DetailResponse(msg=f"获取数据失败: {str(e)}", code=4000)
        assert self.import_field_dict, "'%s' 请配置对应的导入模板字段。" % self.__class__.__name__
        assert self.import_serializer_class, "'%s' 请配置对应的导入序列化器。" % self.__class__.__name__
        # 限制批量更新模板的数据量，避免生成超大文件导致超时
        max_rows = int(request.query_params.get("max_rows") or os.getenv("IMPORT_UPDATE_TEMPLATE_MAX_ROWS", "5000"))
        try:
            total_rows = queryset.count()
        except Exception:
            total_rows = None
        if max_rows and total_rows is not None and total_rows > max_rows:
            return Response(
                {
                    "code": 4000,
                    "msg": f"当前筛选结果 {total_rows} 条，超过模板上限 {max_rows}。请先缩小筛选范围后再下载批量更新模板。",
                    "data": {"total": total_rows, "max_rows": max_rows},
                },
                status=200,
            )
        # 支持 limit/page/page_size 参数，便于前端按筛选条件控制模板行数
        limit_raw = request.query_params.get("limit") or request.query_params.get("page_size")
        page_raw = request.query_params.get("page") or "1"
        try:
            limit = int(limit_raw) if limit_raw else 0
        except Exception:
            limit = 0
        try:
            page = int(page_raw)
        except Exception:
            page = 1
        if limit and page:
            start = max(page - 1, 0) * limit
            end = start + limit
            queryset = queryset[start:end]
        try:
            data = self.import_serializer_class(queryset, many=True, context={"request": request}).data
        except Exception as e:
            return DetailResponse(msg=f"生成模板数据失败: {str(e)}", code=4000)
        # 导出excel 表；先写入 BytesIO 再赋给 response，确保返回正确 xlsx 二进制
        # 仅使用 filename="xxx.xlsx" 避免部分浏览器把 filename* 解析进文件名（如出现 template.xlsx; filename_）
        wb = Workbook()
        ws1 = wb.create_sheet("data", 1)
        ws1.sheet_state = "hidden"
        ws = wb.active
        import_field_dict = {}
        header_data = ["序号","更新主键(勿改)"]
        hidden_header = ["#","id"]
        field_specs = []
        #----设置选项----
        validation_data_dict = {}
        for index, item in enumerate(self.import_field_dict.items()):
            items = list(item)
            key = items[0]
            value = items[1]
            if isinstance(value, dict):
                title = value.get("title")
                header_data.append(str(title) if title is not None else "")
                display_key = value.get('display') or key
                hidden_header.append(display_key)
                field_specs.append({"display": display_key, "choices": value.get("choices", {})})
                choices = value.get("choices", {})
                if choices.get("data"):
                    data_list = []
                    data_list.extend(choices.get("data").keys())
                    validation_data_dict[title] = data_list
                elif choices.get("queryset") and choices.get("values_name"):
                    data_list = choices.get("queryset").values_list(choices.get("values_name"), flat=True)
                    validation_data_dict[title] = list(data_list)
                else:
                    continue
                column_letter = get_column_letter(len(validation_data_dict))
                dv = DataValidation(
                    type="list",
                    formula1=f"{quote_sheetname('data')}!${column_letter}$2:${column_letter}${len(validation_data_dict[title]) + 1}",
                    allow_blank=True,
                )
                ws.add_data_validation(dv)
                dv.add(f"{get_column_letter(index + 3)}2:{get_column_letter(index + 3)}1048576")
            else:
                header_data.append(str(value) if value is not None else "")
                hidden_header.append(key)
                field_specs.append({"display": key, "choices": {}})
        # 添加数据列
        ws1.append(list(validation_data_dict.keys()))
        for index, validation_data in enumerate(validation_data_dict.values()):
            for inx, ele in enumerate(validation_data):
                ws1[f"{get_column_letter(index + 1)}{inx + 2}"] = ele
        #--------
        df_len_max = [self.get_string_len(ele) for ele in header_data]
        row = get_column_letter(len(hidden_header) + 1)
        column = 1
        ws.append(header_data)
        def _map_choice_value(choices, raw_value):
            if raw_value is None or raw_value == "":
                return ""
            if choices.get("data"):
                data_map = choices.get("data") or {}
                for label, value in data_map.items():
                    if raw_value == value:
                        return label
                return raw_value
            if choices.get("queryset") and choices.get("values_name"):
                qs = choices.get("queryset")
                values_name = choices.get("values_name")
                if isinstance(raw_value, (list, tuple)):
                    ids = [v for v in raw_value if v is not None]
                    if not ids:
                        return ""
                    pairs = list(qs.filter(id__in=ids).values_list("id", values_name))
                    id_map = {pid: name for pid, name in pairs}
                    for rid in ids:
                        if rid in id_map:
                            return id_map[rid]
                    return ""
                # 若已是展示值（如 "总所管理-张总"），不要用 id 查，避免 ValueError
                try:
                    int(raw_value)
                    is_id = True
                except (TypeError, ValueError):
                    is_id = False
                if is_id:
                    pair = qs.filter(id=raw_value).values_list(values_name, flat=True).first()
                    return pair if pair is not None else raw_value
                return raw_value
            return raw_value

        for index, results in enumerate(data):
            results_list = []
            for h_index, h_item in enumerate(hidden_header):
                # 跳过序号占位列（由 index+1 写入）
                if h_item == "#":
                    continue
                val = results.get(h_item)
                mapped = val
                for spec in field_specs:
                    if spec["display"] == h_item:
                        mapped = _map_choice_value(spec["choices"], val)
                        break
                if mapped is None or mapped == "":
                    results_list.append("")
                else:
                    results_list.append(mapped)
                # 计算最大列宽度
                if isinstance(mapped, str):
                    result_column_width = self.get_string_len(mapped)
                    if h_index != 0 and result_column_width > df_len_max[h_index]:
                        df_len_max[h_index] = result_column_width
            ws.append([index+1,*results_list])
            column += 1
        # 　更新列宽
        for index, width in enumerate(df_len_max):
            ws.column_dimensions[get_column_letter(index + 1)].width = width
        tab = Table(displayName="Table", ref=f"A1:{row}{column}")  # 名称管理器
        style = TableStyleInfo(
            name="TableStyleLight11",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Access-Control-Expose-Headers"] = "Content-Disposition"
        response["Content-Disposition"] = 'attachment; filename="customer_import_template.xlsx"'
        return response


class ExportSerializerMixin:
    """
    自定义导出功能
    """

    # 导出字段
    export_field_label = []
    # 导出序列化器
    export_serializer_class = None
    # 表格表头最大宽度，默认50个字符
    export_column_width = 50

    def is_number(self, num):
        if isinstance(num, (list, dict)):
            return False
        try:
            float(num)
            return True
        except (ValueError, TypeError):
            pass
        try:
            import unicodedata
            unicodedata.numeric(num)
            return True
        except (TypeError, ValueError):
            pass
        return False

    def get_string_len(self, string):
        """
        获取字符串最大长度；支持 list/dict 等非标量（转为字符串后计算）
        """
        length = 4
        if string is None:
            return length
        if isinstance(string, list):
            string = ", ".join(str(x) for x in string)
        elif isinstance(string, dict):
            string = str(string)
        if self.is_number(string):
            return length
        try:
            for char in str(string):
                length += 2.1 if ord(char) > 256 else 1
        except (TypeError, ValueError):
            return length
        return round(length, 1) if length <= self.export_column_width else self.export_column_width

    def _normalize_export_value(self, val):
        """导出时把 list/dict 转为可写入 Excel 的标量，避免列宽计算和写入异常"""
        if val is None or val == "":
            return ""
        if isinstance(val, list):
            return ", ".join(str(x) for x in val)
        if isinstance(val, dict):
            return str(val)
        return val

    @action(methods=['get'],detail=False)
    def export_data(self, request: Request, *args, **kwargs):
        """
        导出功能
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        queryset = self.filter_queryset(self.get_queryset())
        assert self.export_field_label, "'%s' 请配置对应的导出模板字段。" % self.__class__.__name__
        assert self.export_serializer_class, "'%s' 请配置对应的导出序列化器。" % self.__class__.__name__
        data = self.export_serializer_class(queryset, many=True, context={"request": request}).data
        try:
            async_export_data.delay(
                data,
                str(f"导出{get_verbose_name(queryset)}-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"),
                DownloadCenter.objects.create(creator=request.user, task_name=f'{get_verbose_name(queryset)}数据导出任务', dept_belong_id=request.user.dept_id).pk,
                self.export_field_label
            )
            return SuccessResponse(msg="导入任务已创建，请前往‘下载中心’等待下载")
        except:
            pass
        # 导出excel 表
        response = HttpResponse(content_type="application/msexcel")
        response["Access-Control-Expose-Headers"] = f"Content-Disposition"
        response["content-disposition"] = f'attachment;filename={quote(str(f"导出{get_verbose_name(queryset)}.xlsx"))}'
        wb = Workbook()
        ws = wb.active
        header_data = ["序号", *self.export_field_label.values()]
        hidden_header = ["#", *self.export_field_label.keys()]
        df_len_max = [self.get_string_len(ele) for ele in header_data]
        row = get_column_letter(len(self.export_field_label) + 1)
        column = 1
        ws.append(header_data)
        for index, results in enumerate(data):
            results_list = []
            for h_index, h_item in enumerate(hidden_header):
                for key, val in results.items():
                    if key == h_item:
                        cell_val = self._normalize_export_value(val)
                        results_list.append(cell_val)
                        # 计算最大列宽度（对 list/dict 已归一化为字符串）
                        result_column_width = self.get_string_len(cell_val)
                        if h_index != 0 and result_column_width > df_len_max[h_index]:
                            df_len_max[h_index] = result_column_width
            ws.append([index + 1, *results_list])
            column += 1
        # 　更新列宽
        for index, width in enumerate(df_len_max):
            ws.column_dimensions[get_column_letter(index + 1)].width = width
        tab = Table(displayName="Table", ref=f"A1:{row}{column}")  # 名称管理器
        style = TableStyleInfo(
            name="TableStyleLight11",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(response)
        return response
