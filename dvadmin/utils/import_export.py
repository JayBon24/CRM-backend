# -*- coding: utf-8 -*-
import io
import os
import re
from datetime import datetime
from urllib.parse import urlparse

import openpyxl
from django.conf import settings

from dvadmin.utils.validator import CustomValidationError


def _resolve_excel_source(file_url):
    """
    支持三类导入源：
    1) 本地相对路径：media/files/...
    2) 预览路由：admin-api/system/file/<id>/preview/
    3) 完整URL：http(s)://.../admin-api/system/file/<id>/preview/
    """
    if not file_url:
        raise CustomValidationError("导入文件地址不能为空")

    raw = str(file_url).strip()
    parsed = urlparse(raw)
    path = parsed.path if parsed.scheme else raw
    path = path.lstrip("/")

    # 兼容 DB 存储预览路由
    match = re.search(r"(?:^|/)admin-api/system/file/(\d+)/preview/?$", path)
    if match:
        from dvadmin.system.models import FileList
        file_obj = FileList.objects.filter(id=int(match.group(1))).first()
        if not file_obj:
            raise CustomValidationError("导入文件不存在")

        # db 引擎直接从二进制读取
        if (file_obj.engine or "").lower() == "db":
            if not file_obj.file_blob:
                raise CustomValidationError("导入文件内容为空")
            return io.BytesIO(file_obj.file_blob)

        # 其他引擎优先用本地文件字段
        try:
            if file_obj.url and hasattr(file_obj.url, "path") and os.path.exists(file_obj.url.path):
                return file_obj.url.path
        except Exception:
            pass
        if file_obj.file_url:
            path = str(file_obj.file_url).lstrip("/")

    # 普通本地相对路径
    full_path = os.path.join(settings.BASE_DIR, path)
    if os.path.exists(full_path):
        return full_path

    raise CustomValidationError("导入文件不存在或不可读取")


def get_excel_header(file_url):
    source = _resolve_excel_source(file_url)
    workbook = openpyxl.load_workbook(source, data_only=True)
    table = workbook[workbook.sheetnames[0]]
    header = tuple(table.values)[0] if table.max_row > 0 else tuple()
    workbook.close()
    if hasattr(source, "close"):
        source.close()
    return header


def import_to_data(file_url, field_data, m2m_fields=None):
    """
    读取导入的excel文件
    :param file_url:
    :param field_data: 首行数据源
    :param m2m_fields: 多对多字段
    :return:
    """
    # 读取excel 文件（兼容本地路径与 file_blob）
    source = _resolve_excel_source(file_url)
    workbook = openpyxl.load_workbook(source, data_only=True)
    table = workbook[workbook.sheetnames[0]]
    theader = tuple(table.values)[0] #Excel的表头
    is_update = '更新主键(勿改)' in theader #是否导入更新
    if is_update is False: #不是更新时,删除id列
        field_data.pop('id')
    # 获取参数映射
    validation_data_dict = {}
    for key, value in field_data.items():
        if isinstance(value, dict):
            choices = value.get("choices", {})
            data_dict = {}
            if choices.get("data"):
                for k, v in choices.get("data").items():
                    data_dict[k] = v
            elif choices.get("queryset") and choices.get("values_name"):
                data_list = choices.get("queryset").values(choices.get("values_name"), "id")
                for ele in data_list:
                    data_dict[ele.get(choices.get("values_name"))] = ele.get("id")
            else:
                continue
            validation_data_dict[key] = data_dict
    # 创建一个空列表，存储Excel的数据
    tables = []
    for i, row in enumerate(range(table.max_row)):
        if i == 0:
            continue
        array = {}
        for index, item in enumerate(field_data.items()):
            items = list(item)
            key = items[0]
            values = items[1]
            value_type = 'str'
            if isinstance(values, dict):
                value_type = values.get('type','str')
            cell_value = table.cell(row=row + 1, column=index + 2).value
            if cell_value is None or cell_value=='':
                continue
            elif value_type == 'date':
                print(61, datetime.strptime(str(cell_value), '%Y-%m-%d %H:%M:%S').date())
                try:
                    cell_value = datetime.strptime(str(cell_value), '%Y-%m-%d %H:%M:%S').date()
                except:
                    raise CustomValidationError('日期格式不正确')
            elif value_type == 'datetime':
                cell_value = datetime.strptime(str(cell_value), '%Y-%m-%d %H:%M:%S')
            else:
            # 由于excel导入数字类型后，会出现数字加 .0 的，进行处理
                if type(cell_value) is float and str(cell_value).split(".")[1] == "0":
                    cell_value = int(str(cell_value).split(".")[0])
                elif type(cell_value) is str:
                    cell_value = cell_value.strip(" \t\n\r")
            if key in validation_data_dict:
                array[key] = validation_data_dict.get(key, {}).get(cell_value, None)
                # 选项未匹配时保留原始单元格值，供后端解析多值（如经办人 "张三, 李四"）
                if array[key] is None and cell_value:
                    array[key] = cell_value
                if key in m2m_fields:
                    array[key] = list(
                        filter(
                            lambda x: x,
                            [
                                validation_data_dict.get(key, {}).get(value, None)
                                for value in re.split(r"[，；：|.,;:\s]\s*", cell_value)
                            ],
                        )
                    )
            else:
                array[key] = cell_value
        tables.append(array)
    data = [i for i in tables if len(i) != 0]
    workbook.close()
    if hasattr(source, "close"):
        source.close()
    return data
