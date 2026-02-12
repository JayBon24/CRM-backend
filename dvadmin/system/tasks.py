from hashlib import md5
from io import BytesIO
from datetime import datetime
from time import sleep

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from django.core.files.base import ContentFile

from application.celery import app
from dvadmin.system.models import DownloadCenter

def is_number(num):
    if isinstance(num, (list, dict)):
        return False
    try:
        float(num)
        return True
    except (ValueError, TypeError):
        pass
    try:
        import unicodedata
        unicodedata.numeric(num)
        return True
    except (TypeError, ValueError):
        pass
    return False


def _normalize_export_value(val):
    """导出时把 list/dict 转为可写入 Excel 的标量"""
    if val is None or val == "":
        return ""
    if isinstance(val, list):
        return ", ".join(str(x) for x in val)
    if isinstance(val, dict):
        return str(val)
    return val


def get_string_len(string):
    """
    获取字符串最大长度；支持 list/dict 等非标量（转为字符串后计算）
    """
    length = 4
    if string is None:
        return length
    if isinstance(string, list):
        string = ", ".join(str(x) for x in string)
    elif isinstance(string, dict):
        string = str(string)
    if is_number(string):
        return length
    try:
        for char in str(string):
            length += 2.1 if ord(char) > 256 else 1
    except (TypeError, ValueError):
        return length
    return round(length, 1) if length <= 50 else 50

@app.task
def async_export_data(data: list, filename: str, dcid: int, export_field_label: dict):
    instance = DownloadCenter.objects.get(pk=dcid)
    instance.task_status = 1
    instance.save()
    sleep(2)
    try:
        wb = Workbook()
        ws = wb.active
        header_data = ["序号", *export_field_label.values()]
        hidden_header = ["#", *export_field_label.keys()]
        df_len_max = [get_string_len(ele) for ele in header_data]
        row = get_column_letter(len(export_field_label) + 1)
        column = 1
        ws.append(header_data)
        for index, results in enumerate(data):
            results_list = []
            for h_index, h_item in enumerate(hidden_header):
                for key, val in results.items():
                    if key == h_item:
                        if isinstance(val, datetime):
                            cell_val = val.strftime("%Y-%m-%d %H:%M:%S")
                        else:
                            cell_val = _normalize_export_value(val)
                        results_list.append(cell_val)
                        # 计算最大列宽度（对 list/dict 已归一化为字符串）
                        result_column_width = get_string_len(cell_val)
                        if h_index != 0 and result_column_width > df_len_max[h_index]:
                            df_len_max[h_index] = result_column_width
            ws.append([index + 1, *results_list])
            column += 1
        # 　更新列宽
        for index, width in enumerate(df_len_max):
            ws.column_dimensions[get_column_letter(index + 1)].width = width
        tab = Table(displayName="Table", ref=f"A1:{row}{column}")  # 名称管理器
        style = TableStyleInfo(
            name="TableStyleLight11",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        s = md5()
        while True:
            chunk = stream.read(1024)
            if not chunk:
                break
            s.update(chunk)
        stream.seek(0)
        instance.md5sum = s.hexdigest()
        instance.file_name = filename
        instance.url.save(filename, ContentFile(stream.read()))
        instance.task_status = 2
    except Exception as e:
        instance.task_status = 3
        instance.description = str(e)[:250]
    instance.save()
